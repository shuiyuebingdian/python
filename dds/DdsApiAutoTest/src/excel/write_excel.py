# _*_ coding:utf-8 _*_
import os
import shutil

import cparser
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import RED, GREEN, DARKYELLOW

from src.common.constant import *
from src.config import setting

# --------- 读取conf.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(setting.TEST_CONFIG, encoding=ENCODING)
name = cf.get("tester", "name")


class WriteExcel(object):
    """文件写入数据"""

    def __init__(self, file_name):
        self.file_name = file_name
        if not os.path.exists(self.file_name):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            shutil.copyfile(setting.SOURCE_FILE, setting.TARGET_FILE)
        self.wb = load_workbook(self.file_name)
        # 得到当前可用的worksheet
        self.ws = self.wb.active

    def write_data(self, row_num, value):
        """
        将测试结果写入指定行的RESULT_COL_NUM列
        """
        font_green = Font(name=EXCEL_FONT, color=GREEN, bold=True)
        font_red = Font(name=EXCEL_FONT, color=RED, bold=True)
        align = Alignment(horizontal=EXCEL_ALIGN, vertical=EXCEL_ALIGN)
        # 要写入的结果所在行列坐标
        result_coord = RESULT_COL + str(row_num)
        self.ws.cell(row_num, RESULT_COL_NUM, value)
        if value == PASS:
            self.ws[result_coord].font = font_green
        else:
            self.ws[result_coord].font = font_red
        self.ws[result_coord].alignment = align
        self.wb.save(self.file_name)
